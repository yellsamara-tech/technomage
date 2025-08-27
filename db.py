import os
import asyncpg
import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DB_URL = os.getenv("DATABASE_URL")


# ----- Инициализация базы -----
async def init_db():
    conn = await asyncpg.connect(DB_URL)

    # Таблица пользователей
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id BIGINT PRIMARY KEY,
            full_name TEXT NOT NULL
        )
    """)
    await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS status TEXT;")
    await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS last_update DATE;")
    await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_admin BOOLEAN DEFAULT FALSE;")

    # Таблица истории статусов
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS status_history (
            user_id BIGINT REFERENCES users(id),
            status TEXT,
            status_date DATE,
            PRIMARY KEY(user_id, status_date)
        )
    """)

    await conn.close()


# ----- Добавление пользователя -----
async def add_user(user_id: int, full_name: str, is_admin: bool = False):
    conn = await asyncpg.connect(DB_URL)
    await conn.execute("""
        INSERT INTO users (id, full_name, is_admin) 
        VALUES ($1, $2, $3) 
        ON CONFLICT (id) DO NOTHING
    """, user_id, full_name, is_admin)
    await conn.close()


# ----- Получение пользователя -----
async def get_user(user_id: int):
    conn = await asyncpg.connect(DB_URL)
    row = await conn.fetchrow("SELECT * FROM users WHERE id=$1", user_id)
    await conn.close()
    return row


# ----- Обновление статуса -----
async def update_status(user_id: int, status: str):
    today = date.today()
    conn = await asyncpg.connect(DB_URL)
    await conn.execute("""
        UPDATE users SET status=$1, last_update=$2 WHERE id=$3
    """, status, today, user_id)
    await conn.execute("""
        INSERT INTO status_history(user_id, status, status_date)
        VALUES($1, $2, $3)
        ON CONFLICT(user_id, status_date) DO UPDATE SET status=$2
    """, user_id, status, today)
    await conn.close()


# ----- Получение всех пользователей -----
async def get_all_users():
    conn = await asyncpg.connect(DB_URL)
    rows = await conn.fetch("SELECT * FROM users ORDER BY full_name")
    await conn.close()
    return rows


# ----- Получение истории пользователя -----
async def get_status_history(user_id: int):
    conn = await asyncpg.connect(DB_URL)
    rows = await conn.fetch(
        "SELECT status, status_date FROM status_history WHERE user_id=$1 ORDER BY status_date",
        user_id
    )
    await conn.close()
    return rows


# ----- Поиск пользователей по имени -----
async def find_user_by_name(name: str):
    conn = await asyncpg.connect(DB_URL)
    rows = await conn.fetch(
        "SELECT * FROM users WHERE full_name ILIKE $1", f"%{name}%"
    )
    await conn.close()
    return rows


# ----- Получение админов -----
async def get_admins():
    conn = await asyncpg.connect(DB_URL)
    rows = await conn.fetch("SELECT * FROM users WHERE is_admin=TRUE")
    await conn.close()
    return rows


# ----- Установка статуса напрямую -----
async def set_user_status(user_id: int, status: str):
    await update_status(user_id, status)


# ----- Получение матрицы статусов за месяц -----
async def get_month_matrix(year: int, month: int):
    conn = await asyncpg.connect(DB_URL)

    days_in_month = calendar.monthrange(year, month)[1]

    day_cases = []
    for d in range(1, days_in_month + 1):
        day_cases.append(
            f"MAX(CASE WHEN EXTRACT(DAY FROM sh.status_date) = {d} "
            f"THEN sh.status END) AS d{d:02d}"
        )
    day_cases_sql = ",\n       ".join(day_cases)

    query = f"""
        SELECT u.id, u.full_name,
               {day_cases_sql}
        FROM users u
        LEFT JOIN status_history sh
          ON u.id = sh.user_id
         AND EXTRACT(YEAR FROM sh.status_date) = $1
         AND EXTRACT(MONTH FROM sh.status_date) = $2
        GROUP BY u.id, u.full_name
        ORDER BY u.full_name
    """

    rows = await conn.fetch(query, year, month)
    await conn.close()
    return rows


# ----- Экспорт матрицы в Excel -----
async def export_month_matrix_to_excel(year: int, month: int, filename: str = None):
    rows = await get_month_matrix(year, month)
    days_in_month = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # Заголовки
    headers = ["ID", "ФИО"] + [str(d) for d in range(1, days_in_month + 1)]
    ws.append(headers)

    # Данные
    for row in rows:
        values = [row["id"], row["full_name"]]
        for d in range(1, days_in_month + 1):
            col_name = f"d{d:02d}"
            values.append(row[col_name])
        ws.append(values)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    if not filename:
        filename = f"status_matrix_{year}_{month:02d}.xlsx"

    wb.save(filename)
    return filename
